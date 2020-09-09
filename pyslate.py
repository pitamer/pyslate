# Pyslate
# By Pitamar
# For Awesome Linguists

import datetime
import os
import logging
import re
from tkinter import *
from tkinter.ttk import *

import bs4
import openpyxl
import requests
from translate import Translator


# basic configuration

version_num = '1.27'

original_cwd = os.getcwd()

languages = ['Arabic', 'Persian', 'English']

separators1 = ['.', ',']
separators2 = ['.', ',', None]

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logging.disable(logging.CRITICAL) # <-- Comment out to see logging during run



# functions

def get_sentences_web(source, sep1, sep2):
    """
    Return a list of sentences from a given web source
    One or two sentence separators can be defined
    """
    res = requests.get(source)
    res.raise_for_status()
    
    soup = bs4.BeautifulSoup(res.text, 'html.parser')
    paragraphs = soup.find_all('p')

    logging.debug('Found %s paragraphs.' % (len(paragraphs)))        
    
    sentences = []

    if sep2 == None:
        for p in paragraphs:
            p = p.text.replace('،',',')
            p_sentences = p.split(str(sep1))
            sentences += p_sentences
    else:
        for p in paragraphs:
            p = p.text.replace('،',',')
            p_sentences = p.replace(sep2, sep1).split(str(sep1))
            sentences += p_sentences

    logging.debug('Found %s sentences.' % (len(sentences)))

    return sentences



def clean_sentences_web(sentences):
    """
    Clean sentences of cite notes (like this: [5])
    Return sentences
    """
    cite_note = re.compile(r'\[\d+\]')
    sentences = [re.sub(cite_note, '', s) for s in sentences]
    sentences = [s for s in sentences if len(s) > 0]
    return sentences



def excel_list(lst, sheet, column):
    """
    Export a given list to a given column on a given excel sheet
    Return None
    """
    row = 15
    for i in lst:
        sheet[column+str(row)].value = i
        row += 1
    
    logging.debug('Successfully exported sentences to excel')



def get_translations(sentences, from_lang, to_lang):
    """
    Given list of sentences, translate each to a given language
    Return translations as a list
    """

    translator = Translator(from_lang=from_lang, to_lang=to_lang)
    translations = []

    for s in sentences:
        try:
            trans = translator.translate(s)
        except:
            trans = ''
        translations.append(trans)
        
    return translations
    
    ogging.debug('Successfully translated fros %s to %s' % (from_lang, to_lang))




def save_excel():
    """
    Save and close the excel file open on openpyxl
    Go back to original program directory
    Return None
    """
    os.chdir(folder.get())
    
    stamp = str(datetime.datetime.now().timestamp()).replace('.','_')
    wb.save('output_' + stamp + '.xlsx')
    wb.close()
    
    os.chdir(original_cwd)



def assign_from_lang():
    """
    Determine and return the user's chosen source language
    """
    if src_lang.get() == 'Arabic':
        from_lang = 'ar'
    elif src_lang.get() == 'Persian':
        from_lang = 'fa'
    elif src_lang.get() == 'English':
        from_lang = 'en'
    return from_lang


def pyslate_web():
    """
    Run the whole process of:
    get_sentences_web, clean_sentences_web, get_translations
    excel_list for all lists aquired
    Return 0 for fail, 1 for success
    """
    global wb, sh
    
    wb = openpyxl.load_workbook('format.xlsx')
    sh = wb['פורמט']

    source = entry_text_src.get()

    from_lang = assign_from_lang()
    
    sh['B3'].value = name.get()
    sh['B6'].value = source
    sh['B9'].value = src_lang.get()
    sh['D9'].value = 'תרגום ידני' if (need_heb.get() == 0 and need_eng.get() == 0) else 'תיקון תרגום מכונה'

    update_status('Getting sentences...')
    try:
        sentences = get_sentences_web(source, sep1.get(), sep2.get())
        sentences = clean_sentences_web(sentences)
        excel_list(sentences, sh, 'C')
    except:
        update_status('Sorry, couldn\'t get sentences from the source :(')
        return 0

    if need_heb.get() == True:
        update_status('Translating to Hebrew...')
        try:
            translations = get_translations(sentences, from_lang, 'he')
            excel_list(translations, sh, 'E')
        except:
            update_status('Sorry, wasn\'t able to translate :(')
            save_excel()
            return 0
        
    if need_eng.get() == True:
        update_status('Translating to English...')
        try:
            translations = get_translations(sentences, from_lang, 'en')
            excel_list(translations, sh, 'F')
        except:
            update_status('Sorry, wasn\'t able to translate :(')
            save_excel()
            return 0
        
    save_excel()

    update_status('Your file is ready! :)')
    
    return 1


def get_sentences_xl(sheet):
    """
    Return a list of sentences from a given excel sheet
    """
    sentences = []
    for cell in sheet['C15':'C'+str(len(sheet['C']))]:
        s = str(cell[0].value) if cell[0].value != None else ''
        sentences.append(s)
    return sentences


def get_times_xl(sheet):
    """
    Return a list of times from a given excel sheet
    """
    times = []
    for cell in sheet['G15':'G'+str(len(sheet['G']))]:
        t = str(cell[0].value) if cell[0].value != None else ''
        times.append(t)
    return times


def clean_sentences_xl(sentences):
    """
    Return sentences, cleaned of notes (like this: {PHONE})
    """
    note = re.compile(r'\{.*\}')
    sentences = [re.sub(note, '', s) for s in sentences]
    return sentences


def pyslate_xl(source_sh):
    """
    Run the whole process of:
    get_sentences_xl, clean_sentences_xl, get_translations
    excel_list for all lists aquired
    Return 0 for fail, 1 for success
    """
    global wb, sh
    
    wb = openpyxl.load_workbook('format.xlsx')
    sh = wb['פורמט']

    source = entry_text_src.get()

    from_lang = assign_from_lang()
    
    sh['B3'].value = name.get()
    sh['B6'].value = source
    sh['B9'].value = src_lang.get()
    sh['D9'].value = 'תרגום ידני' if (need_heb.get() == 0 and need_eng.get() == 0) else 'תיקון תרגום מכונה'

    
    update_status('Getting sentences...')
    try:
        sentences = get_sentences_xl(source_sh)
        sentences = clean_sentences_xl(sentences)
        times = get_times_xl(source_sh)
        excel_list(sentences, sh, 'C')
        excel_list(times, sh, 'G')
    except:
        update_status('Sorry, couldn\'t get sentences from the source :(')
        return 0
    
    if need_heb.get() == True:
        update_status('Translating to Hebrew...')
        try:
            translations = get_translations(sentences, from_lang, 'he')
            excel_list(translations, sh, 'E')
        except:
            update_status('Sorry, wasn\'t able to translate :(')
            save_excel()
            return 0
        
    if need_eng.get() == True:
        update_status('Translating to English...')
        try:
            translations = get_translations(sentences, from_lang, 'en')
            excel_list(translations, sh, 'F')
        except:
            update_status('Sorry, wasn\'t able to translate :(')
            save_excel()
            return 0
        
    save_excel()

    update_status('Your file is ready! :)')
    
    return 1
    

def pyslate():
    """
    Determine if the given source is a url or an excel file
    run the appropriate process
    """
    source = entry_text_src.get()
    if source.startswith('http'):
        pyslate_web()
    else:
        source_wb = '%s\%s' % (folder.get(), source)
        if not source.endswith('.xlsx'):
            source_wb += '.xlsx'
        source_wb = openpyxl.load_workbook(source_wb)
        source_sh = source_wb.worksheets[0]
        pyslate_xl(source_sh)


def save_settings():
    """
    Get all the tcl setting variables from their matching widgets
    Assign those variables to the apropriate settings_list index
    Export the updated settings_list to settings.txt (overwrite)
    Eventually update settings_list with load_settings()
    Return None
    """
    global settings_list
    
    settings_list[1]  = folder.get()
    settings_list[3]  = name.get()
    settings_list[5]  = need_heb.get()
    settings_list[7]  = need_eng.get()
    settings_list[9]  = src_lang.get()
    settings_list[11] = sep1.get()
    settings_list[13] = sep2.get()
    
    with open('settings.txt', 'w', encoding="utf8") as writer:
        for li in settings_list:
            writer.write(str(li)+'\n')

    settings_list = load_settings() # <-- settings are updated on saving

    try:
        os.chdir(folder.get())
        os.chdir(original_cwd)
        settings.destroy()
    except:
        entry_save_to.insert(END, '  <-- WRONG PATH! Please fix it')

    
def load_settings():
    """
    Read settings.txt and create a list of settings
    Assign relevant settings_list indexes to apropriate tcl setting variables
    Return settings_list
    """
    with open('settings.txt', 'r', encoding="utf8") as file:
        settings_list = [x[:-1] for x in file.readlines()]

    folder.set(settings_list[1])
    name.set(settings_list[3])
    need_heb.set(settings_list[5])
    need_eng.set(settings_list[7])
    src_lang.set(settings_list[9])
    sep1.set(settings_list[11])
    sep2.set(settings_list[13])

    return settings_list



# front-end

about_message = \
"""
Pyslate %s
Developed by Pitamar
For the best linguists out there

First version released: 18.3.2020
Current version released: 15.7.2020

Pyslate automatically translates all the sentences
found in a given URL or excel file, using Google
Translate. It then exports it all nicely to a new
excel format. Enjoy! :)

Feedback is welcome at: pyslate.dev@gmail.com
(Seriously, please send me feedback!)
""" % (version_num)


def about():
    message = Toplevel(root)
    message.title(' About Pyslate')
    message.iconbitmap('about.ico')
    
    about_label = Label(message, text=about_message)
    button_save = Button(message, text='OK', width=50, command=message.destroy)
    
    about_label.grid(row=0, column=0, padx=std_pad, pady=0)
    button_save.grid(row=1, column=0, padx=std_pad, pady=std_pad)
    

def update_status(status_str):
    """
    Set status (tcl variable, used as textvariable in label_status) to a lot of
    empty tabs, to cover the message that it had before, then change it to the
    message you want to show. Return None
    """
    status.set('\t\t\t\t\t\t\t\t\t\t')
    root.update_idletasks()
    
    status.set(status_str)
    root.update_idletasks()


def settings_window():
    global settings, entry_save_to
    
    settings = Toplevel(root)
    settings.title(' Settings')
    settings.iconbitmap('settings.ico')

    label_save_to = Label(settings, text='Save to:')
    entry_save_to = Entry(settings, width=70, textvariable=folder)

    label_name = Label(settings, text='Name:')
    entry_name = Entry(settings, width=25, textvariable=name)

    check_heb = Checkbutton(settings, text='Hebrew translation', variable=need_heb)

    check_eng = Checkbutton(settings, text='English translation', variable=need_eng)
    
    label_lang = Label(settings, text='Language:')
    combo_lang = Combobox(settings, width=22, values=languages, textvariable=src_lang, state='readonly')

    label_sep1 = Label(settings, text='Sep 1:')
    combo_sep1 = Combobox(settings, width=4, values=separators1, textvariable=sep1, state='readonly')

    label_sep2 = Label(settings, text='Sep 2:')
    combo_sep2 = Combobox(settings, width=4, values=separators2, textvariable=sep2, state='readonly')
    
    button_save = Button(settings, text='Save my settings', width=80, command=save_settings, style='big.TButton')




    label_save_to.grid(row=0, column=0, padx=std_pad, pady=std_pad)
    entry_save_to.grid(row=0, column=1, columnspan=5, padx=std_pad, pady=std_pad)

    label_name.grid(row=1, column=0, padx=std_pad, pady=std_pad)
    entry_name.grid(row=1, column=1, padx=std_pad, pady=std_pad)

    check_heb.grid(row=1, column=2, columnspan=2, padx=std_pad, pady=std_pad)

    check_eng.grid(row=1, column=4, columnspan=2, padx=std_pad, pady=std_pad)

    label_lang.grid(row=2, column=0, padx=std_pad, pady=std_pad)
    combo_lang.grid(row=2, column=1, padx=std_pad, pady=std_pad)

    label_sep1.grid(row=2, column=2, padx=std_pad, pady=std_pad)
    combo_sep1.grid(row=2, column=3, padx=std_pad, pady=std_pad)

    label_sep2.grid(row=2, column=4, padx=std_pad, pady=std_pad)
    combo_sep2.grid(row=2, column=5, padx=std_pad, pady=std_pad)

    button_save.grid(row=3, column=0, columnspan=6, padx=std_pad, pady=std_pad)
    
    # Thanks to the way Tk and tcl handle textvariables and variables,
    # they are automatically inserted to the matching widgets from
    # the settings loaded at the beggining of the run




root = Tk()
root.title(' Pyslate %s' % (version_num))
root.iconbitmap('dict.ico')

std_pad = 8
big_pad = 16
status = StringVar()
status.set('Waiting for instructions')

folder = StringVar()
name = StringVar()
need_heb = IntVar()
need_eng = IntVar()
src_lang = StringVar()
sep1 = StringVar()
sep2 = StringVar()

settings_list = load_settings() # <-- settings are loaded on root startup


label_text_src = Label(root, text='Text Source:')
entry_text_src = Entry(root, width=70)

button_go = Button(root, text='Pyslate it!', width=70, command=pyslate, style='big.TButton')

button_settings = Button(root, text='Settings', width=10, command=settings_window)
button_about = Button(root, text='About', width=10, command=about)

label_status = Label(root, textvariable=status)



label_text_src.grid(row=0, column=0, padx=std_pad, pady=big_pad)
entry_text_src.grid(row=0, column=1, columnspan=2, padx=std_pad, pady=big_pad)

button_go.grid(row=1, column=0, columnspan=2, rowspan=2, padx=std_pad, pady=2)

button_settings.grid(row=1, column=2, padx=0, pady=0)
button_about.grid(row=2, column=2, padx=0, pady=0)

label_status.grid(row=6, column=0, columnspan=3, padx=std_pad, pady=std_pad)



Style().configure('big.TButton', padding=(0, 15))



root.mainloop()
